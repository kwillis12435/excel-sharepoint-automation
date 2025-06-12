#!/usr/bin/env python3
"""
Debug script to examine what header text exists in the calculation sheets.
"""

import os
from openpyxl import load_workbook
from process_study import Config

def debug_sheet_headers():
    """Examine headers in calculation sheets of failing studies."""
    
    base_folder = Config.MONTH_FOLDER
    
    # One of the failing studies
    study_folder = os.path.join(base_folder, "2024013001 (hACVR2b_SEAP_KD_1)")
    results_folder = os.path.join(study_folder, "Results")
    
    if not os.path.exists(results_folder):
        print(f"Results folder not found: {results_folder}")
        return
    
    # Find results file
    results_file = None
    for file in os.listdir(results_folder):
        if file.endswith(('.xlsm', '.xlsx')):
            results_file = os.path.join(results_folder, file)
            break
    
    if not results_file:
        print("No results file found")
        return
    
    print(f"Examining: {results_file}")
    print("=" * 60)
    
    try:
        wb = load_workbook(results_file, data_only=True, read_only=True)
        
        print(f"Available sheets: {wb.sheetnames}")
        print()
        
        # Look at the "Calcs Norm to Ctrl" sheet specifically
        target_sheets = [s for s in wb.sheetnames if "calcs norm" in s.lower()]
        
        for sheet_name in target_sheets:
            print(f"📋 Examining sheet: '{sheet_name}'")
            print("-" * 40)
            
            ws = wb[sheet_name]
            
            # Search for any text that might indicate data sections
            print("🔍 Searching for potential header patterns...")
            
            # Look in the range where relative expression headers usually are
            search_rows = range(120, 140)  # Same range as Config.REL_EXP_SEARCH_ROWS
            
            headers_found = []
            for row in search_rows:
                for col in range(1, 20):  # Check first 20 columns
                    try:
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            cell_text = str(cell_value).strip()
                            if len(cell_text) > 3:  # Skip single characters
                                # Look for potential headers
                                if any(keyword in cell_text.lower() for keyword in [
                                    'expression', 'relative', 'fold', 'change', 'target', 
                                    'gene', 'analysis', 'results', 'summary', 'data'
                                ]):
                                    headers_found.append((row, col, cell_text))
                    except:
                        continue
            
            if headers_found:
                print("✅ Potential headers found:")
                for row, col, text in headers_found:
                    col_letter = chr(ord('A') + col - 1) if col <= 26 else f"Col{col}"
                    print(f"   {col_letter}{row}: '{text}'")
            else:
                print("❌ No relevant headers found in search range")
            
            # Also check the first 50 rows for any obvious headers
            print("\n🔍 Checking first 50 rows for any headers...")
            early_headers = []
            for row in range(1, 51):
                for col in range(1, 10):
                    try:
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            cell_text = str(cell_value).strip()
                            if len(cell_text) > 5 and any(keyword in cell_text.lower() for keyword in [
                                'target', 'gene', 'sample', 'group', 'trigger', 'relative', 'expression'
                            ]):
                                early_headers.append((row, col, cell_text))
                    except:
                        continue
            
            if early_headers:
                print("✅ Early headers found:")
                for row, col, text in early_headers[:10]:  # Show first 10
                    col_letter = chr(ord('A') + col - 1) if col <= 26 else f"Col{col}"
                    print(f"   {col_letter}{row}: '{text}'")
            else:
                print("❌ No early headers found")
            
            print()
        
        wb.close()
        
    except Exception as e:
        print(f"Error examining file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    debug_sheet_headers() 