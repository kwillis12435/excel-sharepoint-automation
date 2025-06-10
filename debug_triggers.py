#!/usr/bin/env python3
"""
Debug script to test trigger extraction from a specific study.
This will help identify where the AC values are being corrupted.
"""

import os
from openpyxl import load_workbook

def debug_trigger_extraction(study_code):
    """Debug trigger extraction for a specific study."""
    
    # Path to the specific study
    base_path = r"C:\Users\kwillis\OneDrive - Arrowhead Pharmaceuticals Inc\Discovery Biology - 2024\01 - 2024"
    
    # First try exact match
    study_folder = os.path.join(base_path, study_code)
    
    if not os.path.exists(study_folder):
        # Try to find folder that starts with the study code
        print(f"Exact folder not found: {study_folder}")
        print("Searching for folders that start with the study code...")
        
        for folder_name in os.listdir(base_path):
            if folder_name.startswith(study_code):
                study_folder = os.path.join(base_path, folder_name)
                print(f"Found matching folder: {folder_name}")
                break
        else:
            print(f"No folder found starting with: {study_code}")
            return
    
    # Check the main study file
    info_file = os.path.join(study_folder, f"{study_code}.xlsm")
    if os.path.exists(info_file):
        print(f"\n=== Debugging {study_code} metadata file ===")
        debug_workbook_triggers(info_file, "Procedure Request Form", "B", 80)
    
    # Check the results file
    results_folder = os.path.join(study_folder, "Results")
    if os.path.exists(results_folder):
        for file in os.listdir(results_folder):
            if file.endswith(".xlsm"):
                results_file = os.path.join(results_folder, file)
                print(f"\n=== Debugging {study_code} results file: {file} ===")
                
                # Try different sheet names
                sheet_names = [
                    "Compiled Indiv. & Grp.",
                    "Compiled Indiv. & Grp",
                    "Compiled Indiv & Grp.",
                    "Compiled Indiv & Grp"
                ]
                
                for sheet_name in sheet_names:
                    try:
                        debug_workbook_triggers(results_file, sheet_name, "B", 130)
                        break
                    except:
                        continue
                break

def debug_workbook_triggers(file_path, sheet_name, column, start_row):
    """Debug trigger extraction from a specific workbook sheet."""
    
    print(f"File: {file_path}")
    print(f"Sheet: {sheet_name}")
    print(f"Column: {column}, Starting row: {start_row}")
    
    # Try both data_only modes
    for data_only in [True, False]:
        print(f"\n--- Testing with data_only={data_only} ---")
        
        try:
            wb = load_workbook(file_path, data_only)                               
            y=data_only, read_only=True
            
            if sheet_name not in wb.sheetnames:
                print(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
                wb.close()
                continue
            
            ws = wb[sheet_name]
            
            # Extract triggers
            triggers = []
            row = start_row
            for i in range(20):  # Check first 20 rows
                cell = ws[f"{column}{row}"]
                value = cell.value
                
                if value is None:
                    break
                
                value_str = str(value).strip()
                triggers.append(value_str)
                
                # Check for corruption
                if 'x' in value_str.lower() and len(value_str) < 10:
                    print(f"  Row {row}: POTENTIAL CORRUPTION: '{value_str}'")
                    print(f"    Cell data type: {getattr(cell, 'data_type', 'unknown')}")
                    print(f"    Cell number format: {getattr(cell, 'number_format', 'unknown')}")
                    if hasattr(cell, 'formula') and cell.formula:
                        print(f"    Cell formula: {cell.formula}")
                else:
                    print(f"  Row {row}: '{value_str}'")
                
                row += 1
            
            print(f"Total triggers found: {len(triggers)}")
            wb.close()
            
        except Exception as e:
            print(f"Error with data_only={data_only}: {e}")

if __name__ == "__main__":
    # Test with the specific study you mentioned
    debug_trigger_extraction("2024010303") 