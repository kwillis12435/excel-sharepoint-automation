import pandas as pd
import os
import re
import json
import csv
from datetime import datetime
import traceback
import logging
from typing import Dict, List, Optional, Tuple, Any
from openpyxl import load_workbook

# ========================== LOGGING SETUP ==========================
def setup_logging(log_file_path: str) -> logging.Logger:
    """
    Set up comprehensive logging for the study processing script.
    Creates both file and console handlers with detailed formatting.
    """
    # Create logger
    logger = logging.getLogger('StudyProcessor')
    logger.setLevel(logging.DEBUG)
    
    # Clear any existing handlers
    logger.handlers.clear()
    
    # Create formatters
    detailed_formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    console_formatter = logging.Formatter(
        '%(levelname)s - %(message)s'
    )
    
    # File handler - logs everything
    file_handler = logging.FileHandler(log_file_path, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(detailed_formatter)
    logger.addHandler(file_handler)
    
    # Console handler - logs INFO and above
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(console_formatter)
    logger.addHandler(console_handler)
    
    return logger

# Global logger instance
logger = None

def init_logger(log_file_path: str):
    """Initialize the global logger instance."""
    global logger
    logger = setup_logging(log_file_path)

class Config:
    """
    Central configuration class containing all constants and settings.
    This makes it easy to modify paths, cell locations, and parameters without 
    hunting through the entire codebase.
        """
    # Main folder containing all study subfolders to process
    MONTH_FOLDER = r"C:\Users\kwillis\OneDrive - Arrowhead Pharmaceuticals Inc\Discovery Biology - 2024\01 - 2024"
    DEBUG = True  # Set to True to see detailed debug output during processing
    
    # Excel sheet names we're looking for in the workbooks
    PROCEDURE_SHEET = "Procedure Request Form"  # Contains study metadata
    RELATIVE_EXPRESSION_SHEETS = [              # Contains the actual expression data
        "Compiled Indiv. & Grp.",
        "Compiled Indiv. & Grp",
        "Compiled Indiv & Grp.",
        "Compiled Indiv & Grp",
        "Calcs Norm to D1 & Ctrl",
        "Calcs Norm to Pre & Control",
        "Calcs Norm to D1 & Ctrl."
    ]
    LAR_SHEET = "LAR Sheet"  # Contains additional metadata
    
    # Specific cell locations in the Procedure Request Form sheet
    STUDY_NAME_CELL = "C14"          # Where study name is stored
    SCREENING_MODEL_CELL = "M6"      # Where screening model is stored  
    STUDY_CODE_CELL = "M12"          # Where study code is stored
    TISSUES_START_ROW = 17           # Row where tissue list starts
    TISSUES_COLUMN = "S"             # Column containing tissue names
    TRIGGERS_START_ROW = 80          # Row where trigger list starts
    TRIGGERS_COLUMN = "B"            # Column containing trigger names
    DOSES_COLUMN = "D"               # Column containing doses (parallel to triggers)
    
    # Settings for relative expression data extraction
    REL_EXP_SEARCH_ROWS = range(120, 135)  # Rows to search for "Relative Expression" header
    TARGET_START_COLUMN = 6                # Column F - where targets start
    TARGET_COLUMN_SPACING = 4              # Targets are spaced 4 columns apart (F, J, N, etc.)
    MAX_TARGETS = 30                       # Safety limit to prevent infinite loops
    MAX_TRIGGERS = 20                      # Safety limit for trigger extraction
    
    # Known dose types to search for
    DOSE_TYPES = {'SQ', 'IV', 'IM', 'intratracheal', 'subcutaneous', 'intravenous', 'intramuscular'}
    
    # Known tissue types to differentiate from gene targets
    TISSUE_TYPES = {
        'lung', 'cerebellum', 'cortex', 'eye', 'liver', 'heart', 'kidney', 'spleen',
        'brain', 'muscle', 'skin', 'blood', 'plasma', 'serum', 'bone', 'fat',
        'adipose', 'pancreas', 'stomach', 'intestine', 'colon', 'bladder',
        'prostate', 'ovary', 'uterus', 'hippocampus', 'cerebral cortex',
        'frontal cortex', 'motor cortex', 'brainstem', 'midbrain', 'spinal cord',
        'thoracic spinal cord', 'lumbar spinal cord', 'cervical spinal cord',
        'skeletal muscle', 'cardiac muscle', 'diaphragm', 'quadriceps',
        'gastrocnemius', 'tibialis anterior', 'retina', 'optic nerve','adrenal gland',
        'sciatic nerve','striatum','lymph node','diaphragm', 'kidney cortex','gastroc','triceps',
        'apex','left ventricle','right ventricle','left atrium','right atrium','medial lobe', 'aorta',
        'rlung','llung','macrophage','tri','gast','gst','hrt','iWAT','thalamus','TSC','Cer','Ctx','Left Lateral Lobe',
        'Right Lateral Lobe', 'Medial Lobe','Teste','bone marrow','bone','quadriceps'
    }

# ========================== UTILITY FUNCTIONS ==========================
def debug_print(*args, **kwargs):
    """
    Print debug messages only when DEBUG mode is enabled.
    This allows us to add debugging throughout the code without cluttering 
    normal output.
    """
    if Config.DEBUG:
        print(*args, **kwargs)
    # Also log to file if logger is available
    if logger:
        message = ' '.join(str(arg) for arg in args)
        logger.debug(message)

def load_known_targets() -> Dict[str, str]:
    """
    Load known gene targets from the alltargets.csv file.
    Returns a dictionary mapping target names to their aliases.
    Also creates reverse mappings for aliases.
    """
    targets_dict = {}
    
    # Try to find the targets file
    possible_paths = [
        "alltargets.csv",
        os.path.join(os.path.dirname(__file__), "alltargets.csv"),
        os.path.join(os.getcwd(), "alltargets.csv")
    ]
    
    targets_file = None
    for path in possible_paths:
        if os.path.exists(path):
            targets_file = path
            break
    
    if not targets_file:
        print("Warning: alltargets.csv not found, using basic target detection")
        if logger:
            logger.warning("alltargets.csv not found, using basic target detection")
        return {}
    
    try:
        with open(targets_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                target = row.get('Target', '').strip()
                alias = row.get('Alias', '').strip()
                
                if target and target != 'null':
                    # Add the main target
                    targets_dict[target.upper()] = target
                    
                    # Add aliases if they exist
                    if alias and alias != 'null':
                        aliases = [a.strip() for a in alias.split(',')]
                        for a in aliases:
                            if a:
                                targets_dict[a.upper()] = target
        
        print(f"Loaded {len(targets_dict)} target names and aliases from {targets_file}")
        if logger:
            logger.info(f"Loaded {len(targets_dict)} target names and aliases from {targets_file}")
        
        return targets_dict
        
    except Exception as e:
        print(f"Error loading targets file: {e}")
        if logger:
            logger.error(f"Error loading targets file: {e}")
        return {}

# Global targets dictionary - loaded once at startup
KNOWN_TARGETS = {}

def init_targets():
    """Initialize the global targets dictionary."""
    global KNOWN_TARGETS
    KNOWN_TARGETS = load_known_targets()

def is_known_target(text: str) -> Tuple[bool, Optional[str]]:
    """
    Check if a string represents a known gene target.
    Handles prefixes like 'r', 'm', 'h' before target names.
    
    Args:
        text: String to check
        
    Returns:
        Tuple of (is_target, canonical_target_name)
    """
    if not text or not KNOWN_TARGETS:
        return False, None
    
    text_clean = str(text).strip()
    text_upper = text_clean.upper()
    
    # First try exact match
    if text_upper in KNOWN_TARGETS:
        return True, KNOWN_TARGETS[text_upper]
    
    # Try with common prefixes (species indicators)
    prefixes = ['R', 'M', 'H', 'C']  # rat, mouse, human, cynomolgus
    for prefix in prefixes:
        if text_upper.startswith(prefix) and len(text_upper) > 1:
            target_without_prefix = text_upper[1:]
            if target_without_prefix in KNOWN_TARGETS:
                canonical_name = KNOWN_TARGETS[target_without_prefix]
                return True, f"{prefix.lower()}{canonical_name}"
    
    return False, None

def is_empty_or_zero(value: Any) -> bool:
    """
    Check if a cell value should be considered empty.
    Excel cells can contain None, empty strings, spaces, or zeros.
    This function standardizes the check across all extraction functions.
    """
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    if value == 0 or str(value).strip() == "0":
        return True  
    return False

def normalize_string(text: str) -> str:
    """
    Remove spaces, special characters, and convert to lowercase.
    Used for fuzzy string matching when trigger names might have 
    slight variations between metadata and results sheets.
    """
    if not text:
        return ""
    return ''.join(c for c in str(text).lower().strip() if c.isalnum())

def format_timepoint(timepoint: str) -> str:
    """
    Ensure timepoint starts with 'D' (e.g., "14" becomes "D14").
    Standardizes timepoint format across different data sources.
    """
    if not timepoint:
        return timepoint
    
    timepoint = str(timepoint).strip()
    if not timepoint.startswith('D') and timepoint and timepoint[0].isdigit():
        return f"D{timepoint}"
    return timepoint

def convert_to_numeric(value: Any) -> str:
    """
    Convert values to consistent numeric format with 4 decimal places.
    Handles various input types (None, strings, numbers) and ensures
    consistent formatting in the final CSV output.
    """
    if value is None:
        return ""
    try:
        float_val = float(str(value).strip())
        return f"{float_val:.4f}" if float_val != 0 else "0.0000"
    except (ValueError, TypeError):
        return str(value).strip()

def extract_dose_from_trigger_name(trigger_name: str) -> Optional[str]:
    """
    Extract dose value from trigger name, specifically looking for mpk values.
    
    Args:
        trigger_name: The trigger name to search for dose information
        
    Returns:
        Extracted dose value or None if not found
    """
    if not trigger_name:
        return None
    
    text = str(trigger_name).lower().strip()
    
    # Look for mpk patterns (e.g., "5mpk", "10 mpk", "2.5mpk")
    mpk_patterns = [
        r'(\d+(?:\.\d+)?)\s*mpk',  # Matches "5mpk", "10 mpk", "2.5mpk"
        r'(\d+(?:\.\d+)?)\s*mg/kg',  # Matches "5mg/kg", "10 mg/kg"
    ]
    
    for pattern in mpk_patterns:
        match = re.search(pattern, text)
        if match:
            dose_value = match.group(1)
            if logger:
                logger.debug(f"Extracted dose '{dose_value} mpk' from trigger name: '{trigger_name}'")
            return f"{dose_value} mpk"
    
    # Look for ug patterns specifically
    ug_patterns = [
        r'(\d+(?:\.\d+)?)\s*(ug|μg)\b',  # Matches "250ug", "5ug", etc.
    ]
    
    for pattern in ug_patterns:
        match = re.search(pattern, text)
        if match:
            dose_value = match.group(1)
            dose_unit = match.group(2)
            extracted_dose = f"{dose_value} ug"
            if logger:
                logger.debug(f"Extracted dose '{extracted_dose}' from trigger name: '{trigger_name}'")
            return extracted_dose
    
    return None

def validate_and_standardize_dose(dose_str: str) -> Tuple[Optional[str], bool]:
    """
    Validate and standardize dose format. Allow common dose formats like ug, mpk, mg/kg, ml/kg.
    Also allow multi-dose formats like "2x5mpk", "2 x 5 mpk".
    Only flag truly problematic formats that need manual review.
    
    Args:
        dose_str: Raw dose string from metadata
        
    Returns:
        Tuple of (standardized_dose, is_flagged)
        - standardized_dose: Clean dose in standard format or None
        - is_flagged: True if dose format is problematic and needs review
    """
    if not dose_str or dose_str.strip() in ['', 'None', '--', 'n/a', 'N/A', 'NA']:
        return None, False
    
    dose_clean = str(dose_str).strip()
    
    # Special case: "0" is a valid dose (zero dose/placebo)
    if dose_clean == "0":
        return "0", False
    
    # Check for valid/acceptable formats including multi-dose schedules
    valid_patterns = [
        r'^(\d+(?:\.\d+)?)\s*(ug|μg)$',              # "250 ug", "5ug"
        r'^(\d+(?:\.\d+)?)\s*mpk$',                  # "5 mpk", "10mpk"
        r'^(\d+(?:\.\d+)?)\s*mg/kg$',                # "3 mg/kg", "10 mg/kg"
        r'^(\d+(?:\.\d+)?)\s*ml/kg$',                # "2 ml/kg", "5 ml/kg"
        r'^(\d+(?:\.\d+)?)\s*mg$',                   # "5 mg", "250 mg"
        r'^(\d+(?:\.\d+)?)\s*ml$',                   # "2 ml", "0.5 ml"
        r'^(\d+(?:\.\d+)?)\s*ng/kg$',                # "500 ng/kg"
        r'^(\d+(?:\.\d+)?)\s*ug/kg$',                # "100 ug/kg"
        r'^(\d+)\s*x\s*(\d+(?:\.\d+)?)\s*mpk$',      # "2x5mpk", "2 x 5 mpk"
        r'^(\d+)\s*x\s*(\d+(?:\.\d+)?)\s*mg/kg$',    # "2x10 mg/kg"
        r'^(\d+)\s*x\s*(\d+(?:\.\d+)?)\s*(ug|μg)$',  # "2x250 ug"
        r'^(\d+)\s*x\s*(\d+(?:\.\d+)?)\s*mg$',       # "2x5 mg"
        r'^(\d+)\s*x\s*(\d+(?:\.\d+)?)\s*ml$',       # "2x1 ml"
    ]
    
    for pattern in valid_patterns:
        match = re.match(pattern, dose_clean.lower())
        if match:
            # For multi-dose formats, preserve the original format
            if 'x' in pattern:
                return dose_clean, False  # Keep original multi-dose format
            else:
                dose_value = match.group(1)
                # Preserve the original unit format
                unit_part = dose_clean[len(dose_value):].strip()
                return f"{dose_value} {unit_part}", False
    
    # Flag only truly problematic formats that need review
    problematic_patterns = [
        r'^\d+$',           # Pure numbers without units (except "0" which we handle above)
        r'[a-zA-Z]+\s*\d+', # Units before numbers like "mpk5"
    ]
    
    for pattern in problematic_patterns:
        if re.search(pattern, dose_clean.lower()):
            if logger:
                logger.warning(f"FLAGGED DOSE: '{dose_clean}' - needs manual review")
            return dose_clean, True  # Return original but flag it
    
    # If it contains numbers and looks like a dose, accept it but don't flag
    if re.search(r'\d', dose_clean):
        # Contains numbers, probably a valid dose format we haven't seen before
        return dose_clean, False
    
    # Unknown format without numbers - flag for review
    if logger:
        logger.warning(f"FLAGGED DOSE: '{dose_clean}' - unknown format")
    return dose_clean, True

def detect_dose_type(text: str) -> Optional[str]:
    """
    Detect dose type (SQ, IV, IM, intratracheal) from text.
    
    Args:
        text: Text to search for dose type indicators
        
    Returns:
        Detected dose type or None if not found
    """
    if not text:
        return None
    
    text_lower = str(text).lower().strip()
    
    # Check for exact matches first
    for dose_type in Config.DOSE_TYPES:
        if dose_type.lower() in text_lower:
            # Return standardized format
            if dose_type.lower() in ['sq', 'subcutaneous']:
                return 'SQ'
            elif dose_type.lower() in ['iv', 'intravenous']:
                return 'IV'
            elif dose_type.lower() in ['im', 'intramuscular']:
                return 'IM'
            elif dose_type.lower() == 'intratracheal':
                return 'Intratracheal'
    
    return None

def is_tissue_name(text: str) -> bool:
    """
    Check if a string represents a tissue name rather than a gene target.
    Only returns True for exact (case-insensitive, stripped) matches to known tissue types.
    """
    if not text:
        return False
    text_lower = str(text).lower().strip()
    # Only allow exact matches
    return text_lower in Config.TISSUE_TYPES


def classify_target_or_tissue(text: str, procedure_tissues: List[str]) -> Tuple[str, Optional[str], Optional[str]]:
    """
    Classify a text string as either a tissue or gene target.
    Uses known targets list with prefix handling for better accuracy.
    Simple rule: If it's a tissue, keep it as tissue. If it's a target, keep it as target.
    No more complex splitting - one item = one classification.
    """
    if not text:
        return 'target', None, None
    text_clean = str(text).strip()
    
    # First check if it's a known gene target (with prefix handling)
    is_target_known, canonical_name = is_known_target(text_clean)
    if is_target_known:
        if Config.DEBUG:
            print(f"    Found known target '{text_clean}' -> '{canonical_name}'")
        return 'target', None, canonical_name
    
    # Check if the entire string is a tissue (exact match to procedure tissues)
    for proc_tissue in procedure_tissues:
        if normalize_string(text_clean) == normalize_string(proc_tissue):
            return 'tissue', text_clean, None
    
    # Check if the entire string is a known tissue type
    if is_tissue_name(text_clean):
        return 'tissue', text_clean, None
    
    # Check if any words in the text are tissue names (but keep the whole thing as tissue)
    words = text_clean.split()
    if len(words) > 1:  # Only check multi-word strings
        # Check for multi-word tissue names first
        all_tissue_types = list(Config.TISSUE_TYPES) + procedure_tissues
        sorted_tissues = sorted(all_tissue_types, key=len, reverse=True)
        
        for tissue_type in sorted_tissues:
            if tissue_type.lower() in text_clean.lower():
                if Config.DEBUG:
                    print(f"    Found tissue phrase '{tissue_type}' in '{text_clean}' - treating whole thing as tissue")
                return 'tissue', text_clean, None
        
        # Check individual words for tissue matches
        for word in words:
            word_clean = word.strip()
            if is_tissue_name(word_clean):
                if Config.DEBUG:
                    print(f"    Found tissue word '{word_clean}' in '{text_clean}' - treating whole thing as tissue")
                return 'tissue', text_clean, None
            
            # Also check against procedure tissues
            for proc_tissue in procedure_tissues:
                if normalize_string(word_clean) == normalize_string(proc_tissue):
                    if Config.DEBUG:
                        print(f"    Found procedure tissue word '{word_clean}' in '{text_clean}' - treating whole thing as tissue")
                    return 'tissue', text_clean, None
    
    # If not a known target or tissue, default to target
    if Config.DEBUG:
        print(f"    Result: TARGET (unknown) - '{text_clean}'")
    return 'target', None, text_clean


def safe_workbook_operation(file_path: str, operation_func, *args, **kwargs):
    """
    Safely open Excel workbooks and ensure they're properly closed.
    Always uses read_only mode unless explicitly disabled.
    ALWAYS uses data_only=True to evaluate formulas properly.
    """
    try:
        # Default to read_only=True unless explicitly set to False
        read_only = kwargs.pop('read_only', True)
        
        # Always use data_only=True to evaluate formulas
        # This is critical for trigger extraction since column B contains formulas
        if logger:
            logger.debug(f"Opening workbook: {file_path} (read_only={read_only}, data_only=True)")
        
        wb = load_workbook(file_path, data_only=True, read_only=read_only)
        result = operation_func(wb, *args)
        wb.close()
        
        if logger:
            logger.debug(f"Successfully processed workbook: {file_path}")
        
        return result
                
    except Exception as e:
        error_msg = f"Error processing {file_path}: {e}"
        print(error_msg)
        if logger:
            logger.error(error_msg)
            logger.error(f"Full traceback: {traceback.format_exc()}")
        if Config.DEBUG:
            traceback.print_exc()
        return None

# ========================== EXCEL DATA EXTRACTION ==========================
class ExcelExtractor:
    """
    Utility class containing reusable methods for extracting data from Excel worksheets.
    This centralizes common patterns like extracting columns, finding text, etc.
    """
    
    @staticmethod
    def extract_column_values(ws, start_row: int, column: str, stop_on_empty: bool = True, max_empty_cells: int = 1) -> List[str]:
        """
        Extract all values from a single column starting at a specific row.
        
        Args:
            ws: Excel worksheet object
            start_row: Row number to start extraction (1-indexed)
            column: Column letter (e.g., 'B', 'S')
            stop_on_empty: Whether to stop when hitting empty cells
            max_empty_cells: Number of consecutive empty cells to tolerate before stopping
            
        Returns:
            List of extracted string values
        """
        values = []
        row = start_row
        empty_count = 0
        
        while row <= ws.max_row:
            # Get cell and handle potential Excel formula issues
            cell = ws[f"{column}{row}"]
            cell_value = cell.value
            
            # Cell value should be properly evaluated with data_only=True
            
            if is_empty_or_zero(cell_value):
                empty_count += 1
                if stop_on_empty and empty_count >= max_empty_cells:
                    break
            else:
                # Reset empty counter when we find a value
                empty_count = 0
                values.append(str(cell_value).strip())
                
            row += 1
            
        return values

    @staticmethod
    def extract_paired_columns(ws, start_row: int, col1: str, col2: str) -> Tuple[List[str], List[str]]:
        """
        Extract values from two parallel columns (e.g., triggers and doses).
        Stops when the first column becomes empty, ensuring both lists have same length.
        
        Args:
            ws: Excel worksheet object
            start_row: Row number to start extraction
            col1: First column letter (e.g., 'B' for triggers)
            col2: Second column letter (e.g., 'D' for doses)
            
        Returns:
            Tuple of (values_from_col1, values_from_col2)
        """
        values1, values2 = [], []
        row = start_row
        while row <= ws.max_row:
            val1 = ws[f"{col1}{row}"].value
            if is_empty_or_zero(val1):
                break
                
            val2 = ws[f"{col2}{row}"].value
            values1.append(str(val1).strip())
            values2.append(val2 if val2 is not None else None)
            row += 1
        return values1, values2

    @staticmethod
    def find_cell_with_text(ws, search_text: str, search_rows: range = None) -> Optional[Tuple[int, int]]:
        """
        Find the first cell containing specific text (case-insensitive).
        
        Args:
            ws: Excel worksheet object
            search_text: Text to search for
            search_rows: Optional range of rows to limit search
            
        Returns:
            Tuple of (row, column) if found, None otherwise
        """
        search_range = search_rows or range(1, min(ws.max_row, 200))
        for row in search_range:
            for col in range(1, min(ws.max_column, 50)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and search_text.lower() in cell_value.lower():
                    return row, col
        return None

    @staticmethod
    def extract_targets_from_row(ws, row: int, procedure_tissues: List[str] = None) -> Tuple[List[str], List[int], List[str], List[str], List[int]]:
        """
        Extract target gene names and tissue names with their column positions from a specific row.
        Now tracks both targets and tissues for data extraction.
        Targets are typically spaced every 4 columns (F, J, N, R, etc.).
        
        Args:
            ws: Excel worksheet object
            row: Row number containing target names
            procedure_tissues: List of tissues from Procedure Request Form for comparison
            
        Returns:
            Tuple of (target_names, target_columns, found_tissues, tissue_names_for_data, tissue_columns_for_data)
        """
        if procedure_tissues is None:
            procedure_tissues = []
            
        targets, target_columns, found_tissues = [], [], []
        tissue_names_for_data, tissue_columns_for_data = [], []  # NEW: Track tissues for data extraction
        col_start = Config.TARGET_START_COLUMN  # Start at column F (6)
        zero_count = 0
        
        while len(targets) + len(found_tissues) < Config.MAX_TARGETS:
            cell_value = ws.cell(row=row, column=col_start).value
            
            if is_empty_or_zero(cell_value):
                zero_count += 1
                if zero_count >= 5:  # Stop after 5 consecutive empty cells
                    break
            else:
                zero_count = 0
            text_clean = str(cell_value).strip()
            
            # Skip obviously invalid target names (Excel errors, pure numbers, etc.)
            if (text_clean.startswith('#') or  # Excel errors like #DIV/0!
                text_clean.replace('.', '').replace('-', '').isdigit() or  # Pure numbers like "1", "0.38"
                len(text_clean) < 2):  # Too short to be meaningful
                print(f"Skipping invalid target name: '{text_clean}'")
                if logger:
                    logger.debug(f"Skipped invalid target name: '{text_clean}'")
            else:
                # Classify as tissue or target (simplified - no more "both")
                classification, tissue_name, target_name = classify_target_or_tissue(
                    text_clean, procedure_tissues
                )
                print(f"Target row cell '{text_clean}': classified as {classification}")
                if logger:
                    logger.debug(f"Target row cell '{text_clean}': classified as {classification}")
                
                if classification == 'tissue':
                    found_tissues.append(tissue_name)
                    # Also add to data extraction lists
                    tissue_names_for_data.append(tissue_name)
                    tissue_columns_for_data.append(col_start)
                    if logger:
                        logger.debug(f"  Added to tissues: '{tissue_name}' (will extract data)")
                else:  # target
                    targets.append(target_name)
                    target_columns.append(col_start)
                    if logger:
                        logger.debug(f"  Added to targets: '{target_name}'")
            
            col_start += Config.TARGET_COLUMN_SPACING  # Move to next target column
        
        return targets, target_columns, found_tissues, tissue_names_for_data, tissue_columns_for_data

# ========================== STUDY METADATA EXTRACTION ==========================
def extract_study_metadata(wb, folder_name: str) -> Dict[str, Any]:
    """
    Extract all metadata from the 'Procedure Request Form' sheet.
    This includes study name, code, screening model, tissues, triggers, doses, and timepoint.
    
    Args:
        wb: Excel workbook object
        folder_name: Name of the study folder (used as fallback for study code)
        
    Returns:
        Dictionary containing all extracted metadata
    """
    if logger:
        logger.info(f"Extracting metadata for study: {folder_name}")
        logger.debug(f"Available sheets: {wb.sheetnames}")
    
    if Config.PROCEDURE_SHEET not in wb.sheetnames:
        if logger:
            logger.warning(f"'{Config.PROCEDURE_SHEET}' sheet not found in {folder_name}, using fallback metadata")
        return _extract_fallback_metadata(folder_name)
    
    ws = wb[Config.PROCEDURE_SHEET]
    
    # Extract basic metadata from specific cells
    study_name = ws[Config.STUDY_NAME_CELL].value
    screening_model = _determine_screening_model(study_name, ws[Config.SCREENING_MODEL_CELL].value)
    study_code = _extract_study_code(ws[Config.STUDY_CODE_CELL].value, folder_name)
    
    # Extract lists of data (tissues, triggers with doses)
    tissues = _extract_unique_tissues(ws)
    
    # Special handling for AAV studies - automatically set tissues to 'serum'
    if study_name and "aav" in str(study_name).lower():
        tissues = ['serum']
        print(f"AAV study detected: '{study_name}' - automatically setting tissues to ['serum']")
        if logger:
            logger.info(f"AAV study detected: '{study_name}' - automatically setting tissues to ['serum']")
    
    raw_triggers, raw_doses = ExcelExtractor.extract_paired_columns(
        ws, Config.TRIGGERS_START_ROW, Config.TRIGGERS_COLUMN, Config.DOSES_COLUMN
    )
    
    # Filter out invalid triggers from metadata but preserve valid biological controls
    triggers, doses = [], []
    for i, trigger in enumerate(raw_triggers):
        trigger_str = str(trigger).strip()
        
        # Skip clearly invalid entries
        if (trigger_str.lower() in ['x', 'n/a', 'blank', 'none', '--', '-', ''] or
            trigger_str.startswith('#') or  # Excel error values
            (trigger_str.replace('.', '').replace(',', '').isdigit() and len(trigger_str) <= 4)):  # Pure small numbers
            if logger:
                logger.debug(f"Filtered out invalid metadata trigger: '{trigger_str}'")
            continue
            
        triggers.append(trigger_str)
        doses.append(raw_doses[i] if i < len(raw_doses) else None)
    
    print(f"Filtered {len(raw_triggers) - len(triggers)} invalid triggers from metadata")
    
    # Debug: Print extracted triggers
    print(f"Extracted {len(triggers)} valid triggers from metadata:")
    if logger:
        logger.info(f"Extracted {len(triggers)} valid triggers from metadata for {folder_name} (filtered {len(raw_triggers) - len(triggers)} invalid)")
    for i, trigger in enumerate(triggers):
        dose = doses[i] if i < len(doses) else None
        print(f"  {i+1}: '{trigger}' -> dose: {dose}")
        if logger:
            logger.debug(f"  Trigger {i+1}: '{trigger}' -> dose: {dose}")
    
    # Create mapping between triggers and their corresponding doses
    trigger_dose_map = _create_trigger_dose_map(triggers, doses)
    
    # Extract timepoint information
    timepoint = _extract_timepoint(ws)
    
    return {
        "study_name": study_name,
        "study_code": study_code,
        "screening_model": screening_model,
        "tissues": tissues,
        "trigger_dose_map": trigger_dose_map,
        "timepoint": timepoint,
    }

def _extract_fallback_metadata(folder_name: str) -> Dict[str, Any]:
    """
    Extract minimal metadata when the procedure sheet is not available.
    Attempts to get study code from folder name using regex pattern.
    """
    study_code = None
    match = re.match(r"(\d{10})", folder_name)  # Look for 10-digit study code
    if match:
        study_code = match.group(1)
    
    return {
        "study_name": None,
        "study_code": study_code,
        "screening_model": None,
        "tissues": [],
        "trigger_dose_map": {},
        "timepoint": None,
    }

def _determine_screening_model(study_name: str, model_cell_value: str) -> str:
    """
    Determine screening model based on study name or specific cell value.
    If study name contains 'aav', automatically set to 'AAV'.
    """
    if study_name and "aav" in str(study_name).lower():
        return "AAV"
    return model_cell_value

def _extract_study_code(cell_value: Any, folder_name: str) -> Optional[str]:
    """
    Extract 10-digit study code from cell value or folder name.
    Validates that the code is exactly 10 digits.
    """
    if cell_value and re.fullmatch(r"\d{10}", str(cell_value)):
        return str(cell_value)
    
    # Fallback to folder name
    match = re.match(r"(\d{10})", folder_name)
    return match.group(1) if match else None

def _extract_unique_tissues(ws) -> List[str]:
    """
    Extract unique tissue types from the tissues column.
    Removes duplicates while preserving order.
    """
    tissues = ExcelExtractor.extract_column_values(
        ws, Config.TISSUES_START_ROW, Config.TISSUES_COLUMN
    )
    return list(dict.fromkeys(tissues))  # Remove duplicates, preserve order

def _create_trigger_dose_map(triggers: List[str], doses: List[Any]) -> Dict[str, Dict[str, Any]]:
    """
    Create a dictionary mapping each trigger to its corresponding dose and detected dose type.
    Ensures both lists are the same length by padding doses with None.
    Now also detects dose types (SQ, IV, IM, intratracheal) and extracts doses from trigger names.
    
    IMPORTANT: Handles duplicate trigger names by making keys unique with position index.
    
    Returns:
        Dictionary mapping trigger_key -> {"dose": dose_value, "dose_type": detected_type}
        where trigger_key is "trigger_name" or "trigger_name_#N" for duplicates
    """
    # Ensure lists are same length
    while len(doses) < len(triggers):
        doses.append(None)
    
    trigger_dose_map = {}
    trigger_name_counts = {}  # Track how many times we've seen each trigger name
    
    for trigger, dose in zip(triggers, doses[:len(triggers)]):
        trigger_name = str(trigger).strip()
        dose_str = str(dose) if dose is not None else ""
        detected_dose_type = detect_dose_type(dose_str)
        
        # Validate and standardize the dose
        standardized_dose, is_flagged = validate_and_standardize_dose(dose_str)
        
        # If no valid dose from metadata, try to extract from trigger name
        final_dose = standardized_dose
        if not standardized_dose:
            extracted_dose = extract_dose_from_trigger_name(trigger_name)
            if extracted_dose:
                final_dose, is_flagged = validate_and_standardize_dose(extracted_dose)
                if logger:
                    logger.info(f"No metadata dose for trigger '{trigger_name}', using extracted dose: '{final_dose}'" + 
                               (f" (FLAGGED)" if is_flagged else ""))
        
        # Create unique trigger key to handle duplicates
        if trigger_name in trigger_name_counts:
            trigger_name_counts[trigger_name] += 1
            trigger_key = f"{trigger_name}_#{trigger_name_counts[trigger_name]}"
            if logger:
                logger.debug(f"Duplicate trigger in metadata: '{trigger_name}' -> using key '{trigger_key}'")
        else:
            trigger_name_counts[trigger_name] = 1
            trigger_key = trigger_name
        
        trigger_dose_map[trigger_key] = {
            "dose": final_dose,
            "dose_type": detected_dose_type,
            "dose_flagged": is_flagged,  # Track if dose needs review
            "original_name": trigger_name  # NEW: Keep original name for matching
        }
        
        if is_flagged:
            print(f"⚠️ DOSE FLAGGED for trigger '{trigger_name}': '{dose_str}' -> '{final_dose}'")
        
        if detected_dose_type:
            debug_print(f"Detected dose type '{detected_dose_type}' for trigger '{trigger_name}': {dose_str}")
        
        if logger:
            logger.debug(f"Trigger dose mapping: '{trigger_key}' (original: '{trigger_name}') -> dose: '{final_dose}', type: '{detected_dose_type}', flagged: {is_flagged}")
    
    print(f"Metadata trigger name distribution: {trigger_name_counts}")
    if logger:
        logger.info(f"Metadata trigger name distribution: {trigger_name_counts}")
    
    return trigger_dose_map

def _extract_timepoint(ws) -> Optional[str]:
    """
    Extract the timepoint value using a systematic approach:
    1. First check known locations for specific study types
    2. Scan column A for timepoint headers or patterns
    3. When a timepoint header is found, search downward to find the LAST value
    4. The last value in a timepoint sequence is the actual timepoint we want
    """
    print("\nExtracting timepoint from worksheet...")
    
    # STRATEGY 1: Check specific known locations for certain study types
    known_locations = [
        (24, 1),   # A24 (common location)
        (16, 4),   # D16 for rIL33_8_Alternaria
        (23, 4),   # D23 for rHDM_Pilot_2
        (55, 1),   # A55 for D154
        (145, 4),  # D145 for mTSHR_7
        (154, 1),   # A154 possibly containing timepoint
        (24, 4)    # D24 possibly containing timepoint
    ]
    
    for row, col in known_locations:
        try:
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_text = str(cell_value).lower()
                match = re.search(r'd(\d+)', cell_text)
                if match and match.group(1):
                    print(f"Found direct timepoint in cell ({row},{col}): D{match.group(1)}")
                    # Continue checking other cells to find the LAST value
                    continue
        except:
            continue
    
    # STRATEGY 2: Find any timepoint header in column A
    timepoint_headers = []
    for row in range(1, min(ws.max_row + 1, 100)):
        try:
            cell_value = ws.cell(row=row, column=1).value
            if not cell_value:
                continue
                
            cell_text = str(cell_value).lower()
            # Check for common timepoint header patterns
            if any(keyword in cell_text for keyword in ['day', 'timepoint', 'sacrifice', 'necropsy']):
                print(f"Found timepoint header in A{row}: {cell_value}")
                timepoint_headers.append(row)
        except:
            continue
    
    # STRATEGY 3: For each header, find the LAST value in that section
    timepoint_values = []
    
    for header_row in timepoint_headers:
        last_value = None
        last_row = None
        empty_count = 0
        
        # Search up to 30 rows below the header
        for row in range(header_row + 1, min(header_row + 30, ws.max_row + 1)):
            try:
                cell_value = ws.cell(row=row, column=1).value
                
                if is_empty_or_zero(cell_value):
                    empty_count += 1
                    if empty_count >= 3:  # Stop after 3 consecutive empty cells
                        break
                else:
                    empty_count = 0
                    cell_text = str(cell_value).strip().lower()
                    
                    # Skip rows that look like headers
                    if any(keyword in cell_text for keyword in ['day', 'timepoint', 'date']):
                        continue
                        
                    # Look for day patterns (D16, D23, etc.)
                    d_match = re.search(r'd(\d+)', cell_text)
                    num_match = re.search(r'^(\d+)$', cell_text)
                    
                    if d_match:
                        last_value = f"D{d_match.group(1)}"
                        last_row = row
                        print(f"Found timepoint in A{row}: {last_value}")
                    elif num_match and not any(skip in cell_text for skip in ['/', '-']):  # Avoid dates
                        last_value = f"D{num_match.group(1)}"
                        last_row = row
                        print(f"Found numeric timepoint in A{row}: {last_value}")
            except:
                continue
        
        if last_value:
            print(f"Final timepoint in section A{last_row}: {last_value}")
            timepoint_values.append((last_row, last_value))
    
    # If we found multiple timepoint values, use the one with highest row number
    # This ensures we get the LAST value in the sheet
    if timepoint_values:
        timepoint_values.sort(key=lambda x: x[0], reverse=True)
        print(f"Selected final timepoint: {timepoint_values[0][1]} (from row A{timepoint_values[0][0]})")
        return timepoint_values[0][1]
    
    # STRATEGY 4: Brute force scan for any D## pattern
    # This is our last resort if the other strategies fail
    d_pattern_rows = []
    
    for row in range(1, min(ws.max_row + 1, 200)):
        try:
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                cell_text = str(cell_value).lower()
                match = re.search(r'd(\d+)', cell_text)
                if match:
                    day_num = int(match.group(1))
                    if 1 <= day_num <= 365:  # Valid day range
                        d_pattern_rows.append((row, f"D{day_num}"))
                        print(f"Found D-pattern in A{row}: D{day_num}")
        except:
            continue
    
    # Get the last D-pattern in the sheet
    if d_pattern_rows:
        d_pattern_rows.sort(key=lambda x: x[0], reverse=True)
        print(f"Selected D-pattern: {d_pattern_rows[0][1]} (from row A{d_pattern_rows[0][0]})")
        return d_pattern_rows[0][1]
    
    # STRATEGY 5: Special case for rIL33_8_Alternaria and similar
    # Known to have timepoint info in specific cells
    special_case_cells = [
        (16, 4), (17, 4), (18, 4),  # D column near row 16
        (23, 4), (24, 4), (25, 4),  # D column near row 23
        (145, 1), (146, 1), (147, 1)  # A column near row 145
    ]
    
    for row, col in special_case_cells:
        try:
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_text = str(cell_value).lower()
                # Very specific pattern matching for these cells
                match = re.search(r'd\s*(\d+)', cell_text)
                if match:
                    timepoint = f"D{match.group(1)}"
                    print(f"Found special case timepoint in ({row},{col}): {timepoint}")
                    return timepoint
        except:
            continue
    
    print("No valid timepoint found after exhaustive search")
    return None

# ========================== RELATIVE EXPRESSION DATA EXTRACTION ==========================
def extract_relative_expression_data(wb, procedure_tissues: List[str] = None) -> Optional[Dict[str, Any]]:
    """
    Extract relative expression data from the results sheet.
    This is the main data we're interested in - target genes vs triggers with expression values.
    
    The data structure looks like:
    - Row with target names (F, J, N, etc.)
    - Rows with trigger names in column B, and corresponding data in columns G,H,I then K,L,M etc.
    
    Args:
        wb: Excel workbook object
        procedure_tissues: List of tissues from Procedure Request Form for comparison
    
    Returns:
        Dictionary with 'targets' list, 'relative_expression_data' nested dict, and 'found_tissues'
    """
    if procedure_tissues is None:
        procedure_tissues = []
    
    if logger:
        logger.info("Extracting relative expression data")
        logger.debug(f"Available sheets: {wb.sheetnames}")
        
    sheet_name = _find_relative_expression_sheet(wb)
    if not sheet_name:
        if logger:
            logger.warning("No relative expression sheet found")
        return None
    
    ws = wb[sheet_name]
    print(f"Using sheet: '{sheet_name}'")
    if logger:
        logger.info(f"Using sheet: '{sheet_name}' for relative expression data")
    
    # Find the data section - handle both standard and calculation sheets
    rel_exp_location = None
    target_row = None
    
    # For "Calcs" sheets, try pattern-based detection instead of header search
    if "calc" in sheet_name.lower():
        print(f"Detected calculation sheet - using pattern-based target detection")
        if logger:
            logger.info(f"Detected calculation sheet '{sheet_name}' - using pattern-based detection")
        
        # For calc sheets, look for the pattern of target names in specific rows
        # Try common target row locations for calc sheets
        potential_target_rows = [15, 16, 17, 18, 19, 20, 25, 30]
        
        for test_row in potential_target_rows:
            # Look for non-empty cells in the target columns that could be gene/target names
            test_targets, test_columns, _, test_tissue_names, test_tissue_columns = ExcelExtractor.extract_targets_from_row(
                ws, test_row, procedure_tissues
            )
            # Accept the row if we find EITHER targets OR tissues
            if test_targets or test_tissue_names:
                target_row = test_row
                print(f"Found data items in row {target_row}: targets={test_targets}, tissues={test_tissue_names}")
                if logger:
                    logger.info(f"Pattern detection found data items in row {target_row}: targets={len(test_targets)}, tissues={len(test_tissue_names)}")
                break
                
        if target_row is None:
            print(f"No target OR tissue pattern found in calculation sheet {sheet_name}")
            if logger:
                logger.warning(f"No target OR tissue pattern found in calculation sheet {sheet_name}")
            return None
            
    else:
        # For standard sheets, use header search
        rel_exp_location = ExcelExtractor.find_cell_with_text(
            ws, "relative expression", Config.REL_EXP_SEARCH_ROWS
        )
        if not rel_exp_location:
            print(f"Relative Expression section not found in sheet {sheet_name}")
            if logger:
                logger.warning(f"Relative Expression section not found in sheet {sheet_name}")
            return None
        
        rel_exp_row, _ = rel_exp_location
        target_row = rel_exp_row + 2
    
    targets, target_columns, found_tissues, tissue_names_for_data, tissue_columns_for_data = ExcelExtractor.extract_targets_from_row(
        ws, target_row, procedure_tissues
    )
    
    # Combine targets and tissues for data extraction
    all_data_names = targets + tissue_names_for_data
    all_data_columns = target_columns + tissue_columns_for_data
    
    # Keep studies even if they only have targets OR only have tissues (be more inclusive)
    if not all_data_names:
        print(f"No targets AND no tissues found for data extraction in row {target_row}")
        if logger:
            logger.warning(f"No targets AND no tissues found for data extraction in row {target_row}")
        return None
    
    # Log what we found
    if targets and tissue_names_for_data:
        print(f"Found both targets ({len(targets)}) and tissues ({len(tissue_names_for_data)}) for data extraction")
    elif targets:
        print(f"Found only targets ({len(targets)}) for data extraction - will leave tissue columns blank")
    elif tissue_names_for_data:
        print(f"Found only tissues ({len(tissue_names_for_data)}) for data extraction - will leave target columns blank")
    
    if logger:
        logger.info(f"Data extraction plan: {len(targets)} gene targets, {len(tissue_names_for_data)} tissue targets, {len(all_data_names)} total items")
    
    # Extract trigger names - adjust based on sheet type
    if "calc" in sheet_name.lower():
        # For calc sheets, triggers are usually 1-2 rows below targets
        trigger_start_row = target_row + 1
        # First try 1 row below
        raw_triggers = ExcelExtractor.extract_column_values(
            ws, trigger_start_row, "B", stop_on_empty=False
        )[:Config.MAX_TRIGGERS]
        
        # If no triggers found, try 2 rows below
        if not raw_triggers or all(is_empty_or_zero(t) for t in raw_triggers):
            trigger_start_row = target_row + 2
            raw_triggers = ExcelExtractor.extract_column_values(
                ws, trigger_start_row, "B", stop_on_empty=False
            )[:Config.MAX_TRIGGERS]
            
        # If still no triggers, try 3 rows below
        if not raw_triggers or all(is_empty_or_zero(t) for t in raw_triggers):
            trigger_start_row = target_row + 3
            raw_triggers = ExcelExtractor.extract_column_values(
                ws, trigger_start_row, "B", stop_on_empty=False
            )[:Config.MAX_TRIGGERS]
    else:
        # For standard sheets, triggers are usually 3 rows below the header
        trigger_start_row = target_row + 3
        raw_triggers = ExcelExtractor.extract_column_values(
            ws, trigger_start_row, "B", stop_on_empty=False
        )[:Config.MAX_TRIGGERS]
    
    # Check if we're getting dose values instead of trigger names
    if raw_triggers:
        dose_like_count = 0
        for trigger in raw_triggers[:5]:  # Check first 5 triggers
            trigger_str = str(trigger).strip().lower()
            # Check if it looks like a dose (contains mpk, ug, mg/kg, etc.)
            if any(unit in trigger_str for unit in ['mpk', 'mg/kg', 'ml/kg', 'ug', 'mg', 'ml']):
                dose_like_count += 1
        
        # If most triggers look like doses, we might be in the wrong location
        if dose_like_count >= len(raw_triggers[:5]) * 0.6:  # 60% threshold
            print(f"⚠️ Warning: Found dose-like values in trigger column: {raw_triggers[:3]}")
            print(f"This suggests triggers might be in a different location or sheet format")
            if logger:
                logger.warning(f"Found dose-like values in trigger column: {raw_triggers[:3]} - possible wrong location")
    
    # Debug output for specific problematic study
    if any('mAdi_110_HDAC11_1' in str(wb.path) for attr in ['path'] if hasattr(wb, attr)):
        print(f"🔍 DEBUG - mAdi_110_HDAC11_1 trigger extraction:")
        print(f"  Raw triggers found: {raw_triggers}")
        print(f"  Target row: {target_row}, trigger start row: {trigger_start_row}")
        print(f"  Sheet name: {sheet_name}")
    
    # Filter out invalid triggers but preserve valid biological controls
    triggers = []
    filtered_count = 0
    for trigger in raw_triggers:
        trigger_str = str(trigger).strip()
        
        # Enhanced filtering - but be more careful about compound codes
        if (trigger_str.lower() in ['x', 'n/a', 'blank', 'none', '--', '-', ''] or
            trigger_str.startswith('#')):  # Excel error values
            if logger:
                logger.debug(f"Filtered out invalid trigger: '{trigger_str}'")
            filtered_count += 1
            continue
        
        # Only filter obvious dose patterns, but preserve compound codes
        is_dose_pattern = False
        trigger_lower = trigger_str.lower()
        
        # Check for dose units
        if any(unit in trigger_lower for unit in ['mpk', 'mg/kg', 'ml/kg', 'ug/kg']):
            # But allow if it also contains letters that suggest it's a compound code
            if not any(trigger_str.upper().startswith(prefix) for prefix in ['AC', 'AS', 'AR', 'AH']):
                is_dose_pattern = True
        
        # Check for pure numbers that look like doses (but allow compound codes with numbers)
        if (trigger_str.replace('.', '').replace(' ', '').isdigit() and 
            len(trigger_str) <= 10 and 
            not trigger_str.startswith(('AC', 'AS', 'AR', 'AH'))):
            is_dose_pattern = True
        
        if is_dose_pattern:
            if logger:
                logger.debug(f"Filtered out dose-like trigger: '{trigger_str}'")
            filtered_count += 1
            continue
            
        triggers.append(trigger_str)
    
    if filtered_count > 0:
        print(f"Filtered {filtered_count} invalid/dose-like triggers, kept {len(triggers)} triggers")
        if logger:
            logger.info(f"Filtered {filtered_count} invalid/dose-like triggers from {len(raw_triggers)} raw triggers, kept {len(triggers)} triggers")
    
    # If we don't have many triggers from results sheet, try to use metadata triggers instead
    if len(triggers) < 3 and procedure_tissues:
        print(f"⚠️ Only found {len(triggers)} triggers in results sheet, attempting to match with metadata triggers")
        if logger:
            logger.warning(f"Only found {len(triggers)} triggers in results sheet, attempting metadata matching")
        
        # This will be implemented in the calling function where we have access to metadata
        # For now, continue with what we found
    
    print(f"Found targets: {targets}")
    print(f"Found tissues for data extraction: {tissue_names_for_data}")
    print(f"Found {len(triggers)} triggers in relative expression data:")
    if logger:
        logger.info(f"Found {len(targets)} gene targets, {len(tissue_names_for_data)} tissue targets, and {len(triggers)} triggers in relative expression data")
        logger.debug(f"Gene targets: {targets}")
        logger.debug(f"Tissue targets: {tissue_names_for_data}")
    for i, trigger in enumerate(triggers):
        print(f"  {i+1}: '{trigger}'")
        if logger:
            logger.debug(f"  Trigger {i+1}: '{trigger}'")
    
    if found_tissues:
        print(f"Found tissues in target row: {found_tissues}")
        if logger:
            logger.info(f"Found tissues in target row: {found_tissues}")
    
    # Extract the actual expression data for each trigger-target combination (including tissues)
    print(f"\n🔍 EXTRACTING DATA:")
    print(f"  Triggers to process: {len(triggers)}")
    print(f"  Data items (targets + tissues): {len(all_data_names)}")
    print(f"  Expected trigger-target combinations: {len(triggers)} × {len(all_data_names)} = {len(triggers) * len(all_data_names)}")
    if logger:
        logger.info(f"Data extraction: {len(triggers)} triggers × {len(all_data_names)} items = {len(triggers) * len(all_data_names)} expected combinations")
    
    triggers_data = _extract_trigger_target_data(ws, triggers, all_data_names, all_data_columns, trigger_start_row)
    
    # Count actual combinations found
    actual_combinations = 0
    for trigger_data in triggers_data.values():
        actual_combinations += len(trigger_data)
    
    print(f"  Actual combinations extracted: {actual_combinations}")
    if logger:
        logger.info(f"Actually extracted {actual_combinations} trigger-target combinations")
    
    # Remove triggers with no data
    clean_triggers_data = {k: v for k, v in triggers_data.items() if v}
    
    # Final count after cleaning
    final_combinations = 0
    for trigger_data in clean_triggers_data.values():
        final_combinations += len(trigger_data)
    
    print(f"  Final combinations after cleaning: {final_combinations}")
    if logger:
        logger.info(f"Final {final_combinations} combinations after removing empty triggers")
    
    return {
        "targets": targets,
        "tissue_targets": tissue_names_for_data,  # NEW: Tissues that have expression data
        "relative_expression_data": clean_triggers_data,
        "found_tissues": found_tissues,
        "all_data_items": all_data_names,  # NEW: Combined list of all items with data
        "trigger_start_row": trigger_start_row,  # NEW: For potential re-processing
        "raw_triggers_found": raw_triggers  # NEW: For debugging
    }

def _find_relative_expression_sheet(wb) -> Optional[str]:
    """
    Find the sheet containing relative expression data from the list of possible names.
    Tries exact matches first, then case-insensitive keyword matching.
    """
    # Try exact matches first
    for sheet_name in Config.RELATIVE_EXPRESSION_SHEETS:
        if sheet_name in wb.sheetnames:
            return sheet_name
    
    # Try case-insensitive and keyword matching
    for sheet in wb.sheetnames:
        sheet_lower = sheet.lower()
        if (("compiled" in sheet_lower and ("indiv" in sheet_lower or "grp" in sheet_lower)) or
            ("calcs" in sheet_lower and "norm" in sheet_lower and "ctrl" in sheet_lower)):
            return sheet
    
    return None

def _extract_trigger_target_data(ws, triggers: List[str], targets: List[str], 
                                target_columns: List[int], trigger_start_row: int) -> Dict[str, Dict[str, Dict[str, Any]]]:
    """
    Extract the actual expression values for each trigger-target combination.
    
    For each target at column X, the data is in columns X+1, X+2, X+3:
    - X+1: relative expression value
    - X+2: low confidence interval  
    - X+3: high confidence interval
    
    IMPORTANT: Handles duplicate trigger names by making keys unique with position index.
    
    Args:
        ws: Excel worksheet
        triggers: List of trigger names
        targets: List of target gene names
        target_columns: List of column numbers where targets are located
        trigger_start_row: Row where trigger data begins
        
    Returns:
        Nested dictionary: {trigger_key: {target: {rel_exp, low, high}}}
        where trigger_key is "trigger_name" or "trigger_name_#N" for duplicates
    """
    triggers_data = {}
    combinations_extracted = 0
    combinations_skipped = 0
    trigger_name_counts = {}  # Track how many times we've seen each trigger name
    
    if logger:
        logger.debug(f"Starting data extraction from row {trigger_start_row} for {len(triggers)} triggers and {len(targets)} targets")
        logger.debug(f"Target columns: {target_columns}")
    
    for trigger_idx, trigger in enumerate(triggers):
        if is_empty_or_zero(trigger):
            if logger:
                logger.debug(f"Skipping empty trigger at index {trigger_idx}")
            continue
        
        # Additional safety check to filter out invalid triggers
        trigger_str = str(trigger).strip()
        if (trigger_str.lower() in ['x', 'n/a', 'blank', 'none', '--', '-', ''] or
            trigger_str.startswith('#') or  # Excel error values
            (trigger_str.replace('.', '').replace(',', '').isdigit() and len(trigger_str) <= 4)):  # Pure small numbers
            if logger:
                logger.debug(f"Skipping invalid trigger: '{trigger_str}'")
            continue
            
        # Create unique trigger key to handle duplicates
        trigger_name = trigger_str
        if trigger_name in trigger_name_counts:
            trigger_name_counts[trigger_name] += 1
            trigger_key = f"{trigger_name}_#{trigger_name_counts[trigger_name]}"
            if logger:
                logger.debug(f"Duplicate trigger detected: '{trigger_name}' -> using key '{trigger_key}'")
        else:
            trigger_name_counts[trigger_name] = 1
            trigger_key = trigger_name
            
        trigger_row = trigger_start_row + trigger_idx
        triggers_data[trigger_key] = {}
        
        if logger:
            logger.debug(f"Processing trigger '{trigger_key}' (original: '{trigger_name}', row {trigger_row})")
        
        for target_idx, target in enumerate(targets):
            base_col = target_columns[target_idx]
            
            # Extract the three data values for this trigger-target combination
            values = {
                "rel_exp": ws.cell(row=trigger_row, column=base_col + 1).value,  # Relative expression
                "low": ws.cell(row=trigger_row, column=base_col + 2).value,      # Low CI
                "high": ws.cell(row=trigger_row, column=base_col + 3).value      # High CI
            }
            
            # Debug: Show what we're extracting
            if logger:
                logger.debug(f"  Target '{target}' (col {base_col}): rel_exp={values['rel_exp']}, low={values['low']}, high={values['high']}")
            
            # Skip combinations with no data (all None values)
            if all(v is None for v in values.values()):
                combinations_skipped += 1
                if logger:
                    logger.debug(f"    Skipped: all values are None for {trigger_key} + {target}")
                continue
            
            # Keep combinations that have at least some data
            triggers_data[trigger_key][target] = values
            combinations_extracted += 1
            
            debug_print(f"  ✓ {trigger_key} + {target}: {values}")
            if logger:
                logger.debug(f"    ✓ Extracted data for {trigger_key} + {target}")
    
    if logger:
        logger.info(f"Data extraction completed: {combinations_extracted} combinations extracted, {combinations_skipped} skipped")
        logger.info(f"Trigger name distribution: {trigger_name_counts}")
    
    print(f"  📊 Data extraction summary: {combinations_extracted} extracted, {combinations_skipped} skipped")
    print(f"  🔄 Trigger name counts: {trigger_name_counts}")
    
    return triggers_data

# ========================== STRING MATCHING UTILITIES ==========================
class StringMatcher:
    """
    Utility class for finding the best match between strings.
    Used to match trigger names between metadata and results sheets
    when they might have slight variations.
    """
    
    @staticmethod
    def find_best_match(target: str, candidates: List[str]) -> Optional[str]:
        """
        Find the best matching string using multiple strategies:
        1. Exact match (case-insensitive)
        2. Prefix match (target is at start of candidate)
        3. Normalized match (alphanumeric characters only)
        
        Args:
            target: String to find a match for
            candidates: List of possible matches
            
        Returns:
            Best matching candidate string, or None if no match found
        """
        if not target or not candidates:
            return None
        
        target_clean = target.lower().strip()
        
        # Strategy 1: Exact match
        for candidate in candidates:
            if candidate.lower().strip() == target_clean:
                return candidate
        
        # Strategy 2: Prefix match
        for candidate in candidates:
            if candidate.lower().strip().startswith(target_clean):
                return candidate
        
        # Strategy 3: Normalized match (alphanumeric only)
        target_norm = normalize_string(target)
        for candidate in candidates:
            if normalize_string(candidate).startswith(target_norm):
                return candidate
        
        return None

# ========================== STUDY PROCESSING ==========================
def process_study_folder(study_folder: str) -> Optional[Dict[str, Any]]:
    """
    Process a single study folder and extract all available data.
    
    Each study folder should contain:
    - {folder_name}.xlsm (metadata file)
    - Results/{results_file}.xlsm (results file)
    
    Args:
        study_folder: Path to the study folder
        
    Returns:
        Dictionary containing all extracted data, or None if no data found
    """
    folder_name = os.path.basename(study_folder)
    info_file = os.path.join(study_folder, f"{folder_name}.xlsm")
    results_folder = os.path.join(study_folder, "Results")

    print(f"\nProcessing study: {folder_name}")
    if logger:
        logger.info(f"Starting processing of study: {folder_name}")
        logger.debug(f"Study folder path: {study_folder}")
        logger.debug(f"Expected info file: {info_file}")
        logger.debug(f"Expected results folder: {results_folder}")
    
    study_data = {}

    # Extract metadata from the main study file
    if os.path.exists(info_file):
        if logger:
            logger.debug(f"Found info file: {info_file}")
        metadata = safe_workbook_operation(info_file, extract_study_metadata, folder_name)
        if metadata:
            study_data.update(metadata)
            print("Extracted metadata fields:")
            if logger:
                logger.info(f"Successfully extracted metadata for {folder_name}")
            for k, v in metadata.items():
                print(f"  {k}: {v}")
                if logger:
                    logger.debug(f"  Metadata {k}: {v}")
        else:
            if logger:
                logger.error(f"Failed to extract metadata from {info_file}")
    else:
        print(f"Info file not found: {info_file}")
        if logger:
            logger.warning(f"Info file not found: {info_file}")
    
    # Extract data from the results file
    results_file = _find_results_file(results_folder)
    if results_file:
        if logger:
            logger.debug(f"Found results file: {results_file}")
    else:
        if logger:
            logger.warning(f"No results file found in {results_folder}")
    
    if results_file:
        # Extract LAR sheet data (additional metadata)
        lar_data = _extract_lar_data(results_file)
        if lar_data:
            study_data["lar_data"] = lar_data
            print(f"Extracted LAR data: {lar_data}")
        
        # Extract the main relative expression data
        procedure_tissues = study_data.get("tissues", [])
        rel_exp_data = safe_workbook_operation(
            results_file, 
            extract_relative_expression_data,
            procedure_tissues
        )
        
        if rel_exp_data:
            # Check if we got problematic triggers (dose-like values) and have metadata triggers available
            results_triggers = list(rel_exp_data.get('relative_expression_data', {}).keys())
            metadata_triggers = list(study_data.get('trigger_dose_map', {}).keys())
            
            print(f"🔄 Comparing trigger lists:")
            print(f"  Metadata triggers ({len(metadata_triggers)}): {metadata_triggers}")
            print(f"  Results triggers ({len(results_triggers)}): {results_triggers}")
            
            # NEW LOGIC: Use the trigger list that has more items (longer list)
            if len(metadata_triggers) > len(results_triggers) and 'raw_triggers_found' in rel_exp_data:
                print(f"🔄 Metadata has more triggers ({len(metadata_triggers)} vs {len(results_triggers)}), attempting to use metadata triggers...")
                
                if logger:
                    logger.info(f"Using longer trigger list: metadata has {len(metadata_triggers)} vs results {len(results_triggers)}")
                
                # Try to re-extract with better trigger matching
                enhanced_rel_exp_data = safe_workbook_operation(
                    results_file,
                    _extract_relative_expression_with_metadata_triggers,
                    procedure_tissues,
                    metadata_triggers,
                    rel_exp_data.get('trigger_start_row'),
                    rel_exp_data.get('targets', []),
                    rel_exp_data.get('tissue_targets', [])
                )
                
                if enhanced_rel_exp_data and len(enhanced_rel_exp_data.get('relative_expression_data', {})) > len(results_triggers):
                    print(f"✅ Enhanced matching found {len(enhanced_rel_exp_data.get('relative_expression_data', {}))} triggers vs {len(results_triggers)} original")
                    rel_exp_data = enhanced_rel_exp_data
                    if logger:
                        logger.info(f"Enhanced trigger matching successful: {len(enhanced_rel_exp_data.get('relative_expression_data', {}))} triggers found")
                else:
                    print(f"⚠️ Enhanced matching failed or didn't improve trigger count, keeping original results")
                    if logger:
                        logger.warning(f"Enhanced trigger matching failed to improve results")
            elif len(results_triggers) >= len(metadata_triggers):
                print(f"✅ Results triggers are equal or longer ({len(results_triggers)} vs {len(metadata_triggers)}), keeping results triggers")
                if logger:
                    logger.info(f"Keeping results triggers: {len(results_triggers)} vs metadata {len(metadata_triggers)}")
            else:
                print(f"⚠️ No enhanced matching attempted (no raw triggers data available)")
                if logger:
                    logger.warning(f"Enhanced matching not available (missing raw_triggers_found)")
            
            study_data["relative_expression"] = rel_exp_data
            print(f"Extracted relative expression data:")
            print(f"  - Targets: {len(rel_exp_data.get('targets', []))} targets")
            print(f"  - Triggers: {len(rel_exp_data.get('relative_expression_data', {}))} triggers")
            # Print a sample of the data structure
            if rel_exp_data.get('relative_expression_data'):
                first_trigger = next(iter(rel_exp_data['relative_expression_data']))
                print(f"  - Sample data structure for trigger '{first_trigger}':")
                if rel_exp_data['relative_expression_data'][first_trigger]:
                    first_target = next(iter(rel_exp_data['relative_expression_data'][first_trigger]))
                    print(f"    {first_target}: {rel_exp_data['relative_expression_data'][first_trigger][first_target]}")
                else:
                    print("    No target data found")
        else:
            print("  No relative expression data found")
    
    return study_data if study_data else None

def _find_results_file(results_folder: str) -> Optional[str]:
    """
    Find the first .xlsm file in the Results folder.
    There should typically be only one results file per study.
    """
    if not os.path.exists(results_folder):
        return None
    
    for file in os.listdir(results_folder):
        if file.endswith(".xlsm"):
            return os.path.join(results_folder, file)
    
    return None

def _extract_lar_data(results_file: str) -> Optional[Dict[str, str]]:
    """
    Extract additional metadata from the LAR Sheet.
    Searches for keyword-value pairs in the first 20 rows.
    """
    try:
        if Config.LAR_SHEET not in pd.ExcelFile(results_file).sheet_names:
            return None
        
        df = pd.read_excel(results_file, sheet_name=Config.LAR_SHEET, header=None)
        
        fields = {}
        for i, row in df.iterrows():
            if i >= 20:  # Limit search to first 20 rows
                break
            for col in range(len(row)):
                cell = str(row[col]).strip().lower()
                # Look for cells containing these keywords
                if any(keyword in cell for keyword in ["trigger", "dose", "tissue", "timepoint"]):
                    field_name = next(k for k in ["trigger", "dose", "tissue", "timepoint"] if k in cell)
                    field_value = str(row[col + 1]).strip() if col + 1 < len(row) else ""
                    if field_value != "nan":
                        fields[field_name] = field_value
        
        return fields if fields else None
    except Exception as e:
        print(f"Error extracting LAR data: {e}")
        return None

# ========================== CSV EXPORT ==========================
def export_to_csv(all_study_data: List[Dict[str, Any]], output_path: str):
    """
    Export all study data to a CSV file in a standardized format.
    
    The CSV will have one row per trigger-target combination with columns:
    - study_name, study_code, screening_model, gene_target, trigger, dose, dose_type,
      timepoint, tissue, avg_rel_exp, avg_rel_exp_lsd, avg_rel_exp_hsd
    """
    if logger:
        logger.info(f"Starting CSV export to: {output_path}")
        logger.info(f"Processing {len(all_study_data)} studies for CSV export")
    
    header = [
        "study_name", "study_code", "screening_model", "gene_target", "item_type", "trigger", 
        "dose", "dose_type", "timepoint", "tissue", "avg_rel_exp", "avg_rel_exp_lsd", "avg_rel_exp_hsd"
    ]
    
    csv_rows = [header]
    stats = {
        "studies_processed": 0, 
        "studies_with_data": 0, 
        "studies_excluded": 0,
        "total_rows": 0,
        "excluded_studies": []
    }
    
    for study in all_study_data:
        stats["studies_processed"] += 1
        study_name = study.get('study_name', 'Unknown')
        study_code = study.get('study_code', 'Unknown')
        
        if logger:
            logger.debug(f"Processing study for CSV: {study_name} ({study_code})")
        
        rows_added = _process_study_for_csv(study, csv_rows)
        if rows_added > 0:
            stats["studies_with_data"] += 1
            stats["total_rows"] += rows_added
        else:
            stats["studies_excluded"] += 1
            stats["excluded_studies"].append({
                "name": study_name,
                "code": study_code,
                "reason": "No valid data rows generated"
            })
    
    # Write CSV file
    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        csv.writer(csvfile).writerows(csv_rows)
    
    if logger:
        logger.info(f"CSV file written successfully: {output_path}")
        logger.info("="*60)
        logger.info("CSV EXPORT SUMMARY")
        logger.info("="*60)
        logger.info(f"Studies processed for CSV: {stats['studies_processed']}")
        logger.info(f"Studies included in CSV: {stats['studies_with_data']}")
        logger.info(f"Studies excluded from CSV: {stats['studies_excluded']}")
        logger.info(f"Total data rows in CSV: {stats['total_rows']}")
        
        if stats["excluded_studies"]:
            logger.warning("STUDIES EXCLUDED FROM CSV:")
            for excluded in stats["excluded_studies"]:
                logger.warning(f"  - {excluded['name']} ({excluded['code']}): {excluded['reason']}")
    
    _print_export_summary(stats, output_path)

def _process_study_for_csv(study: Dict[str, Any], csv_rows: List[List[str]]) -> int:
    """
    Process a single study for CSV export.
    Groups all rows for a study by tissue (primary) and then by target (secondary).
    If tissue is missing, groups by target.
    """
    study_name = study.get('study_name', 'Unknown')
    study_code = study.get('study_code', 'Unknown')
    
    if "relative_expression" not in study:
        exclusion_reason = "No relative expression data found"
        print(f"Skipping study with no relative expression data: {study_name}")
        if logger:
            logger.warning(f"EXCLUDED FROM CSV: {study_name} ({study_code}) - {exclusion_reason}")
        return 0

    print(f"\nProcessing CSV data for study: {study_name}")
    if logger:
        logger.info(f"Processing CSV data for study: {study_name} ({study_code})")
    
    study_info = _extract_study_info_for_csv(study)
    print(f"  Study info: {study_info}")
    if logger:
        logger.debug(f"Study info for CSV: {study_info}")

    rel_exp_data = study["relative_expression"].get("relative_expression_data", {})
    if not rel_exp_data:
        exclusion_reason = "Relative expression data structure is empty"
        print(f"  No relative expression data found in study")
        if logger:
            logger.warning(f"EXCLUDED FROM CSV: {study_name} ({study_code}) - {exclusion_reason}")
        return 0

    print(f"  Found {len(rel_exp_data)} triggers in relative expression data")
    print(f"  Trigger names in relative expression: {list(rel_exp_data.keys())}")
    print(f"  Trigger names in metadata: {list(study_info['trigger_dose_map'].keys())}")
    
    if logger:
        logger.debug(f"Found {len(rel_exp_data)} triggers in relative expression data")
        logger.debug(f"Trigger names in relative expression: {list(rel_exp_data.keys())}")
        logger.debug(f"Trigger names in metadata: {list(study_info['trigger_dose_map'].keys())}")

        # Collect all rows for this study
    study_rows = []
    
    # Get lists of gene targets and tissue targets for proper classification
    gene_targets = study.get("relative_expression", {}).get("targets", [])
    tissue_targets = study.get("relative_expression", {}).get("tissue_targets", [])
    
    for trigger_key in rel_exp_data.keys():
            # Extract original trigger name from trigger key (handle duplicates)
            # trigger_key might be "saline" or "saline_#2" for duplicates
            if "_#" in trigger_key:
                trigger_name = trigger_key.rsplit("_#", 1)[0]  # Remove "_#N" suffix
            else:
                trigger_name = trigger_key
            
            # Get dose info from metadata mapping - try exact trigger_key first, then original name
            trigger_info = None
            
            # First try exact match with trigger_key (for duplicates like "Saline_#2")
            if trigger_key in study_info["trigger_dose_map"]:
                trigger_info = study_info["trigger_dose_map"][trigger_key]
            else:
                # Fall back to matching by original name
                for metadata_key, metadata_info in study_info["trigger_dose_map"].items():
                    if metadata_info.get("original_name", metadata_key) == trigger_name:
                        trigger_info = metadata_info
                        break
                
                # Last resort: use trigger name directly
                if not trigger_info:
                    trigger_info = study_info["trigger_dose_map"].get(trigger_name, {"dose": "", "dose_type": ""})
            
            # Try to extract dose from trigger name if not available in mapping
            extracted_dose = extract_dose_from_trigger_name(trigger_name)
            if extracted_dose and not trigger_info.get("dose"):
                trigger_info = {
                    "dose": extracted_dose,
                    "dose_type": trigger_info.get("dose_type", "")
                }
                if logger:
                    logger.info(f"Using extracted dose '{extracted_dose}' for trigger '{trigger_name}'")
            elif trigger_info.get("dose") and extracted_dose:
                # Both sources available - log for comparison
                if logger:
                    logger.debug(f"Trigger '{trigger_name}': metadata dose='{trigger_info.get('dose')}', extracted dose='{extracted_dose}'")
            
            trigger_data = rel_exp_data[trigger_key]
            
            if logger:
                logger.debug(f"Processing trigger key '{trigger_key}' (name: '{trigger_name}') with {len(trigger_data)} items, final dose: '{trigger_info.get('dose', 'None')}'")
            
            for item_name, value in trigger_data.items():
                # Determine if this item is a gene target or tissue target
                if item_name in gene_targets:
                    item_type = "gene_target"
                elif item_name in tissue_targets:
                    item_type = "tissue_target"
                else:
                    item_type = "unknown"  # Fallback
                
                if isinstance(value, dict):
                    row = _create_csv_row(
                        study_info, trigger_name, trigger_info, item_name, item_type,  # Use trigger_name, not trigger_key
                        {"rel_exp": value.get("rel_exp"), "low": value.get("low"), "high": value.get("high")}
                    )
                else:
                    row = _create_csv_row(
                        study_info, trigger_name, trigger_info, item_name, item_type,  # Use trigger_name, not trigger_key
                        {"rel_exp": value, "low": None, "high": None}
                    )
                
                # Skip rows with empty/corrupted trigger names
                if row and row[5]:  # row[5] is the trigger column
                    study_rows.append(row)
                    
                    if logger:
                        logger.debug(f"Added CSV row: {trigger_name} (key: {trigger_key}) -> {item_name} ({item_type}) = {value}, dose = {trigger_info.get('dose', 'None')}")
                else:
                    if logger:
                        logger.debug(f"Skipped CSV row with empty/corrupted trigger: {trigger_name} (key: {trigger_key}) -> {item_name}")

    # Determine index for tissue and target columns in the row
    tissue_idx = 9  # 0-based index for 'tissue' in the row (shifted due to new item_type column)
    target_idx = 3  # 0-based index for 'gene_target' in the row
    item_type_idx = 4  # 0-based index for 'item_type' in the row

    # Sort rows: by tissue (if present), then by item type, then by target
    def sort_key(row):
        tissue = row[tissue_idx] or "ZZZ"  # Put missing tissues at the end
        item_type = row[item_type_idx] or "ZZZ"
        target = row[target_idx] or "ZZZ"
        return (tissue, item_type, target)

    study_rows.sort(key=sort_key)

    # Append sorted rows to the main CSV
    for row in study_rows:
        csv_rows.append(row)

    print(f"  Added {len(study_rows)} rows to CSV for study {study_name}")
    if logger:
        logger.info(f"INCLUDED IN CSV: {study_name} ({study_code}) - Added {len(study_rows)} rows")
    return len(study_rows)

def _extract_study_info_for_csv(study: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract and format study information for CSV export.
    """
    timepoint = study.get("timepoint", "")
    if timepoint and not timepoint.startswith('D') and timepoint.strip().isdigit():
        timepoint = f"D{timepoint.strip()}"
    
    tissue = ""
    if study.get("tissues"):
        tissue = study["tissues"][0]
    elif "lar_data" in study and "tissue" in study["lar_data"]:
        tissue = study["lar_data"]["tissue"]
    
    if not tissue and "relative_expression" in study:
        found_tissues = study["relative_expression"].get("found_tissues", [])
        if found_tissues:
            tissue = found_tissues[0]
    
    return {
        "study_name": study.get("study_name", ""),
        "study_code": f"'{study.get('study_code', '')}'" if study.get("study_code") else "",
        "screening_model": study.get("screening_model", ""),
        "trigger_dose_map": study.get("trigger_dose_map", {}),
        "timepoint": timepoint,
        "tissue": tissue
    }

def _create_csv_row(study_info: Dict[str, Any], trigger: str, trigger_info: Dict[str, Any], 
                   target: str, item_type: str, values: Dict[str, Any]) -> List[str]:
    """
    Create a single CSV row for one trigger-target combination.
    
    NEW LOGIC: If item is a tissue_target, put it in tissue column and leave gene_target empty.
               If item is a gene_target, put it in gene_target column and use study tissue.
    
    CSV columns (in order):
    1. study_name: Name of the study
    2. study_code: 10-digit code (quoted to prevent Excel issues)
    3. screening_model: Type of screening used
    4. gene_target: Name of the target gene (empty if tissue_target)
    5. item_type: Type of item (gene_target or tissue_target)
    6. trigger: Name of the trigger (e.g., siRNA sequence) - CLEANED
    7. dose: Dose amount for this trigger
    8. dose_type: Detected dose type (SQ, IV, IM, Intratracheal)
    9. timepoint: Time point of measurement (e.g., D14)
    10. tissue: Tissue type (from study metadata OR the tissue_target name)
    11. avg_rel_exp: Average relative expression value
    12. avg_rel_exp_lsd: Lower standard deviation
    13. avg_rel_exp_hsd: Higher standard deviation
    
    All numeric values are formatted to 4 decimal places.
    """
    # Handle the case where values is directly a float (new format)
    if not isinstance(values, dict):
        rel_exp = values
        low = None
        high = None
    else:
        rel_exp = values.get("rel_exp")
        low = values.get("low")
        high = values.get("high")
    
    # Clean the trigger name and handle dose extraction
    existing_dose = trigger_info.get("dose", "")
    has_existing_dose = bool(existing_dose and existing_dose.strip())
    
    cleaned_trigger, extracted_dose = clean_trigger_name(trigger, has_existing_dose)
    
    # Use extracted dose if no existing dose is available
    final_dose = existing_dose
    if not has_existing_dose and extracted_dose:
        final_dose = extracted_dose
        if logger:
            logger.debug(f"Using extracted dose '{extracted_dose}' for cleaned trigger '{cleaned_trigger}'")
    
    # NEW LOGIC: Proper tissue vs target placement
    if item_type == "tissue_target":
        # If this is a tissue target, put it in tissue column, leave gene_target empty
        gene_target_col = ""
        tissue_col = target  # The tissue name from results file
    else:
        # If this is a gene target, put it in gene_target column, use study tissue
        gene_target_col = target
        tissue_col = study_info["tissue"]  # Use tissue from study metadata
    
    # Flag if both columns would be empty (shouldn't happen but safety check)
    if not gene_target_col and not tissue_col:
        print(f"⚠️ WARNING: Both gene_target and tissue columns empty for {target} ({item_type})")
        if logger:
            logger.warning(f"Both gene_target and tissue columns empty for {target} ({item_type})")
        
    return [
        study_info["study_name"],
        study_info["study_code"],
        study_info["screening_model"],
        gene_target_col,                       # gene_target (empty for tissue_targets)
        item_type,                             # item_type (gene_target or tissue_target)
        cleaned_trigger,                       # trigger (CLEANED - no doses/volumes/admin details)
        str(final_dose),                       # dose (from metadata or extracted from trigger)
        str(trigger_info.get("dose_type", "")), # dose_type (detected)
        study_info["timepoint"],
        tissue_col,                            # tissue (study tissue OR tissue_target name)
        convert_to_numeric(rel_exp),           # avg_rel_exp 
        convert_to_numeric(low),               # avg_rel_exp_lsd
        convert_to_numeric(high)               # avg_rel_exp_hsd
    ]

def _print_export_summary(stats: Dict[str, Any], output_path: str):
    """Print summary statistics after CSV export."""
    print(f"\nExport Summary:")
    print(f"- Total studies processed: {stats['studies_processed']}")
    print(f"- Studies included in CSV: {stats['studies_with_data']}")
    print(f"- Studies excluded from CSV: {stats.get('studies_excluded', 0)}")
    print(f"- Total data rows: {stats['total_rows']}")
    print(f"- Output file: {output_path}")
    
    if stats.get('excluded_studies'):
        print(f"\nExcluded studies:")
        for excluded in stats['excluded_studies']:
            print(f"  - {excluded['name']} ({excluded['code']}): {excluded['reason']}")

# ========================== MAIN EXECUTION ==========================
def main():
    """
    Main execution function.
    """
    # Set up logging first
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_output_dir = os.path.dirname(Config.MONTH_FOLDER)
    month_name = os.path.basename(Config.MONTH_FOLDER).split(' ')[0]
    log_file_path = os.path.join(base_output_dir, f"study_processing_log_{month_name}_{timestamp}.log")
    
    init_logger(log_file_path)
    
    # Initialize target detection system
    init_targets()
    
    logger.info("="*80)
    logger.info("STUDY PROCESSING SESSION STARTED")
    logger.info("="*80)
    logger.info(f"Processing folder: {Config.MONTH_FOLDER}")
    logger.info(f"Log file: {log_file_path}")
    logger.info(f"Debug mode: {Config.DEBUG}")
    
    study_folders = [
        os.path.join(Config.MONTH_FOLDER, name)
        for name in os.listdir(Config.MONTH_FOLDER)
        if os.path.isdir(os.path.join(Config.MONTH_FOLDER, name))
    ]
    
    if not study_folders:
        error_msg = "No studies found in the month folder."
        print(error_msg)
        logger.error(error_msg)
        return

    print(f"Processing {len(study_folders)} study folders")
    logger.info(f"Found {len(study_folders)} study folders to process")
    logger.debug(f"Study folders: {[os.path.basename(f) for f in study_folders]}")
    
    # Track processing statistics
    processing_stats = {
        "total_folders": len(study_folders),
        "successful_processing": 0,
        "failed_processing": 0,
        "included_in_csv": 0,
        "excluded_from_csv": 0,
        "failed_folders": [],
        "excluded_studies": []
    }
    
    all_study_data = []
    for i, study_folder in enumerate(study_folders, 1):
        folder_name = os.path.basename(study_folder)
        logger.info(f"[{i}/{len(study_folders)}] Processing: {folder_name}")
        
        try:
            study_data = process_study_folder(study_folder)
            if study_data:
                all_study_data.append(study_data)
                processing_stats["successful_processing"] += 1
                logger.info(f"Successfully processed: {folder_name}")
            else:
                processing_stats["failed_processing"] += 1
                processing_stats["failed_folders"].append(folder_name)
                logger.error(f"Failed to extract any data from: {folder_name}")
        except Exception as e:
            processing_stats["failed_processing"] += 1
            processing_stats["failed_folders"].append(folder_name)
            error_msg = f"Exception processing {folder_name}: {e}"
            print(f"ERROR: {error_msg}")
            logger.error(error_msg)
            logger.error(f"Full traceback: {traceback.format_exc()}")

    # Log processing summary
    logger.info("="*80)
    logger.info("PROCESSING SUMMARY")
    logger.info("="*80)
    logger.info(f"Total folders found: {processing_stats['total_folders']}")
    logger.info(f"Successfully processed: {processing_stats['successful_processing']}")
    logger.info(f"Failed processing: {processing_stats['failed_processing']}")
    
    if processing_stats["failed_folders"]:
        logger.warning(f"Failed folders: {processing_stats['failed_folders']}")

    # Save JSON output
    json_timestamp = datetime.now().strftime("%Y%m%d")
    json_output_path = os.path.join(base_output_dir, f"study_metadata_{month_name}_{json_timestamp}.json")
    with open(json_output_path, "w", encoding="utf-8") as f:
        json.dump(all_study_data, f, indent=2, ensure_ascii=False)
    print(f"\nWrote study metadata to {json_output_path}")
    logger.info(f"Wrote study metadata to {json_output_path}")
    
    # Export to CSV with detailed logging
    csv_output_path = os.path.join(base_output_dir, f"study_data_{month_name}_{json_timestamp}.csv")
    logger.info("="*80)
    logger.info("CSV EXPORT STARTING")
    logger.info("="*80)
    
    export_to_csv(all_study_data, csv_output_path)
    
    # Create final summary report
    _create_final_summary_report(processing_stats, all_study_data, log_file_path)
    
    # Final summary
    logger.info("="*80)
    logger.info("SESSION COMPLETED")
    logger.info("="*80)
    logger.info(f"Log file saved to: {log_file_path}")
    print(f"\nDetailed log saved to: {log_file_path}")

def _create_final_summary_report(processing_stats: Dict[str, Any], all_study_data: List[Dict[str, Any]], log_file_path: str):
    """Create a comprehensive summary report at the end of the log file."""
    if not logger:
        return
    
    logger.info("="*80)
    logger.info("FINAL PROCESSING REPORT")
    logger.info("="*80)
    
    # Overall statistics
    logger.info("OVERALL STATISTICS:")
    logger.info(f"  Total study folders found: {processing_stats['total_folders']}")
    logger.info(f"  Successfully processed: {processing_stats['successful_processing']}")
    logger.info(f"  Failed to process: {processing_stats['failed_processing']}")
    logger.info(f"  Success rate: {processing_stats['successful_processing']/processing_stats['total_folders']*100:.1f}%")
    
    # Failed folders
    if processing_stats["failed_folders"]:
        logger.warning("FOLDERS THAT FAILED PROCESSING:")
        for folder in processing_stats["failed_folders"]:
            logger.warning(f"  - {folder}")
    
    # Studies with data breakdown
    studies_with_metadata = sum(1 for study in all_study_data if any(k in study for k in ['study_name', 'study_code', 'tissues']))
    studies_with_rel_exp = sum(1 for study in all_study_data if 'relative_expression' in study)
    studies_with_both = sum(1 for study in all_study_data if 'relative_expression' in study and any(k in study for k in ['study_name', 'study_code']))
    
    logger.info("DATA EXTRACTION BREAKDOWN:")
    logger.info(f"  Studies with metadata: {studies_with_metadata}")
    logger.info(f"  Studies with relative expression data: {studies_with_rel_exp}")
    logger.info(f"  Studies with both metadata and expression data: {studies_with_both}")
    
    # Detailed study breakdown
    logger.info("DETAILED STUDY BREAKDOWN:")
    for study in all_study_data:
        study_name = study.get('study_name', 'Unknown')
        study_code = study.get('study_code', 'Unknown')
        
        has_metadata = any(k in study for k in ['study_name', 'study_code', 'tissues', 'trigger_dose_map'])
        has_rel_exp = 'relative_expression' in study
        
        if has_rel_exp:
            rel_exp_data = study['relative_expression'].get('relative_expression_data', {})
            num_triggers = len(rel_exp_data)
            num_targets = len(study['relative_expression'].get('targets', []))
            total_data_points = sum(len(trigger_data) for trigger_data in rel_exp_data.values())
        else:
            num_triggers = num_targets = total_data_points = 0
        
        status = []
        if has_metadata:
            status.append("metadata")
        if has_rel_exp:
            status.append(f"rel_exp({num_triggers}t×{num_targets}g={total_data_points}pts)")
        
        status_str = ", ".join(status) if status else "no_data"
        logger.info(f"  {study_name} ({study_code}): {status_str}")
    
    logger.info("="*80)

def _extract_relative_expression_with_metadata_triggers(wb, procedure_tissues: List[str], 
                                                     metadata_triggers: List[str], 
                                                     trigger_start_row: int,
                                                     existing_targets: List[str],
                                                     existing_tissue_targets: List[str]) -> Optional[Dict[str, Any]]:
    """
    Enhanced relative expression extraction that uses metadata triggers to match with results data.
    NEW APPROACH: Use positional mapping when name matching fails - map metadata triggers 
    sequentially to data rows that have actual expression values.
    
    Args:
        wb: Excel workbook object
        procedure_tissues: List of tissues from procedure sheet
        metadata_triggers: List of trigger names from metadata (info file)
        trigger_start_row: Row where trigger data starts in results sheet
        existing_targets: Already found gene targets
        existing_tissue_targets: Already found tissue targets
        
    Returns:
        Dictionary with enhanced trigger matching or None if failed
    """
    if logger:
        logger.info(f"Enhanced trigger matching: using {len(metadata_triggers)} metadata triggers")
    
    sheet_name = _find_relative_expression_sheet(wb)
    if not sheet_name:
        return None
    
    ws = wb[sheet_name]
    
    # Use existing target information
    all_data_names = existing_targets + existing_tissue_targets
    
    # Find target columns again (we need this for data extraction)
    target_columns = []
    tissue_columns_for_data = []
    col_start = Config.TARGET_START_COLUMN
    
    for i in range(len(existing_targets)):
        target_columns.append(col_start)
        col_start += Config.TARGET_COLUMN_SPACING
        
    for i in range(len(existing_tissue_targets)):
        tissue_columns_for_data.append(col_start)
        col_start += Config.TARGET_COLUMN_SPACING
    
    all_data_columns = target_columns + tissue_columns_for_data
    
    # Extract all data rows that have actual expression values
    data_rows_with_values = []
    max_trigger_rows = min(len(metadata_triggers) + 10, Config.MAX_TRIGGERS)  # Check more rows
    
    for row_offset in range(max_trigger_rows):
        row = trigger_start_row + row_offset
        if row > ws.max_row:
            break
            
        # Check if this row has any actual data values
        has_data = False
        data_values = []
        
        for col in all_data_columns:
            values = {
                "rel_exp": ws.cell(row=row, column=col + 1).value,
                "low": ws.cell(row=row, column=col + 2).value,
                "high": ws.cell(row=row, column=col + 3).value
            }
            data_values.append(values)
            
            # Check if this column has any non-null data
            if not all(v is None for v in values.values()):
                has_data = True
        
        if has_data:
            row_data = {
                'row': row,
                'trigger_cell': ws.cell(row=row, column=2).value,  # Column B
                'data_values': data_values
            }
            data_rows_with_values.append(row_data)
    
    print(f"  Found {len(data_rows_with_values)} data rows with actual expression values")
    if logger:
        logger.debug(f"Found {len(data_rows_with_values)} data rows with expression values for trigger matching")
    
    # NEW STRATEGY: Use positional mapping - map metadata triggers to data rows sequentially
    triggers_data = {}
    mapped_count = 0
    
    # Map metadata triggers to data rows (one-to-one up to the number of available data rows)
    for i, meta_trigger in enumerate(metadata_triggers):
        if i < len(data_rows_with_values):
            row_data = data_rows_with_values[i]
            trigger_data = {}
            
            # Extract data for all targets/tissues
            for j, target_name in enumerate(all_data_names):
                if j < len(row_data['data_values']):
                    values = row_data['data_values'][j]
                    
                    # Only include if there's actual data
                    if not all(v is None for v in values.values()):
                        trigger_data[target_name] = values
            
            if trigger_data:  # Only add if we found data
                triggers_data[meta_trigger] = trigger_data
                mapped_count += 1
                
                trigger_cell_value = str(row_data['trigger_cell'] or '').strip()
                print(f"  Mapped metadata trigger '{meta_trigger}' to row {row_data['row']} (results cell: '{trigger_cell_value}')")
                if logger:
                    logger.debug(f"Positionally mapped metadata trigger '{meta_trigger}' to row {row_data['row']} (results cell: '{trigger_cell_value}')")
    
    print(f"  Successfully mapped {mapped_count}/{len(metadata_triggers)} metadata triggers to data rows")
    if logger:
        logger.info(f"Positional mapping result: {mapped_count}/{len(metadata_triggers)} triggers mapped")
    
    # Accept any positive mapping result when using the longer list strategy
    if mapped_count > 0:
        return {
            "targets": existing_targets,
            "tissue_targets": existing_tissue_targets,
            "relative_expression_data": triggers_data,
            "found_tissues": [],  # Keep empty since we're reusing existing data
            "all_data_items": all_data_names,
            "enhanced_matching": True,  # Flag to indicate this used enhanced matching
            "positional_mapping": True  # Flag to indicate positional mapping was used
        }
    
    return None

def clean_trigger_name(trigger_name: str, has_existing_dose: bool = False) -> Tuple[str, Optional[str]]:
    """
    Clean trigger names by removing unnecessary information like volumes, doses, and administration details.
    
    Examples:
    - "250uL  HDM 5ug (D1, 3)" -> "HDM" (and extract "5ug" if no existing dose)
    - "200uL  Saline NA (D1, 3)" -> "Saline"
    - "AC005120 2x5mpk (D1, 3)" -> "AC005120" (and extract "2x5mpk" if no existing dose)
    - "AC00008 + saline" -> "AC00008 + saline"
    - "PBS" -> "PBS"
    
    Args:
        trigger_name: Raw trigger name from Excel
        has_existing_dose: Whether there's already a dose in the metadata
        
    Returns:
        Tuple of (cleaned_trigger_name, extracted_dose_if_needed)
    """
    if not trigger_name:
        return "", None
    
    original_name = str(trigger_name).strip()
    
    # First, check for corrupted/invalid trigger names that should be filtered out
    corrupted_patterns = [
        r'^AC\d+\s+\d+\s*x\s*$',           # "AC006365 4 x"
        r'^AC\d+/kg$',                     # "AC007163/kg"
        r'^[A-Z]-[a-zA-Z0-9]+/[a-zA-Z0-9]+/[a-zA-Z]+$',  # "B-hIL11/hIL11RA/mL"
        r'^\d+\s*x\s*$',                   # Just "4 x"
        r'^/kg$',                          # Just "/kg"
        r'^/[a-zA-Z]+$',                   # Just "/mL", "/ug", etc.
        r'^\d+\s*[a-zA-Z]*\s*$',           # Pure numbers with optional units
    ]
    
    for pattern in corrupted_patterns:
        if re.match(pattern, original_name, re.IGNORECASE):
            if logger:
                logger.debug(f"Filtered out corrupted trigger name: '{original_name}'")
            return "", None
    
    # Remove parentheses and everything inside them (D1, 3), (D1, 14), etc.
    cleaned = re.sub(r'\([^)]*\)', '', original_name).strip()
    
    # Remove volume measurements at the beginning (250uL, 200uL, etc.)
    cleaned = re.sub(r'^\d+\s*u?[lL]\s+', '', cleaned).strip()
    
    # Extract dose information before removing it (if no existing dose)
    extracted_dose = None
    if not has_existing_dose:
        extracted_dose = extract_dose_from_trigger_name(cleaned)
    
    # For AC numbers with combinations (like "AC00008 + saline"), preserve the whole thing
    if re.search(r'AC\d+.*[+&].*', cleaned, re.IGNORECASE):
        # This is a compound combination, keep it as-is but remove dose info
        dose_patterns = [
            r'\s+\d+(?:\.\d+)?\s*(?:ug|μg|mg|ml)\b',           # "5ug", "250 mg"
            r'\s+\d+(?:\.\d+)?\s*(?:mpk|mg/kg|ml/kg)\b',       # "5mpk", "10 mg/kg"
            r'\s+\d+\s*x\s*\d+(?:\.\d+)?\s*(?:ug|μg|mg|ml|mpk|mg/kg)\b',  # "2x5mpk", "2x250ug"
            r'\s+NA\b',                                         # "NA" (not applicable)
        ]
        
        for pattern in dose_patterns:
            cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE).strip()
    else:
        # For single compounds, remove dose information
        dose_patterns = [
            r'\s+\d+(?:\.\d+)?\s*(?:ug|μg|mg|ml)\b',           # "5ug", "250 mg"
            r'\s+\d+(?:\.\d+)?\s*(?:mpk|mg/kg|ml/kg)\b',       # "5mpk", "10 mg/kg"
            r'\s+\d+\s*x\s*\d+(?:\.\d+)?\s*(?:ug|μg|mg|ml|mpk|mg/kg)\b',  # "2x5mpk", "2x250ug"
            r'\s+NA\b',                                         # "NA" (not applicable)
        ]
        
        for pattern in dose_patterns:
            cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE).strip()
    
    # Clean up extra whitespace
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    
    # If we ended up with an empty string, use a fallback
    if not cleaned:
        # Try to extract just the core compound/trigger name
        words = original_name.split()
        for word in words:
            # Look for AC numbers, common trigger names, etc.
            if (re.match(r'^AC\d+', word, re.IGNORECASE) or 
                word.upper() in ['SALINE', 'HDM', 'PBS', 'VEHICLE', 'CONTROL', 'ACSF']):
                cleaned = word
                break
        
        # Last resort - use first meaningful word that's not obviously corrupted
        if not cleaned and words:
            for word in words:
                if (len(word) >= 2 and 
                    not re.match(r'^\d+$', word) and  # Not pure numbers
                    not word.endswith('/kg') and      # Not dose fragments
                    not word.endswith('/mL')):        # Not unit fragments
                    cleaned = word
                    break
    
    # Final validation - if it's still corrupted, return empty
    if cleaned and any(re.match(pattern, cleaned, re.IGNORECASE) for pattern in corrupted_patterns):
        if logger:
            logger.debug(f"Final filter caught corrupted trigger: '{cleaned}' from '{original_name}'")
        return "", None
    
    if logger:
        if cleaned != original_name:
            logger.debug(f"Cleaned trigger name: '{original_name}' -> '{cleaned}'" + 
                        (f" (extracted dose: '{extracted_dose}')" if extracted_dose else ""))
    
    return cleaned, extracted_dose

if __name__ == "__main__":
    main()